from flask import Flask, render_template, request, jsonify, session
from flask_cors import CORS
from config import Config
from models import db, Settings, Campaign, EmailTemplate, JobLead, ActivityLog, ManufacturingICPCampaign, ManufacturingLead
from services.google_search import GoogleSearchService
from services.apollo_api import ApolloAPIService
from services.email_generator import EmailGenerator
from services.email_sender import EmailSender
from services.sheets_logger import SheetsLogger
from services.scheduler import CampaignScheduler
from services.job_parser import JobParserService
from services.ai_lead_scorer import AILeadScorer
from datetime import datetime
from urllib.parse import urlparse
import os
import json

app = Flask(__name__)
app.config.from_object(Config)
CORS(app)

# Initialize database
db.init_app(app)

# Initialize services (will be configured from settings)
email_generator = EmailGenerator()
campaign_scheduler = CampaignScheduler()
ai_lead_scorer = AILeadScorer()

def get_setting(key, default=None):
    """Get a setting from database or return default"""
    setting = Settings.query.filter_by(key=key).first()
    return setting.value if setting else default

def save_setting(key, value):
    """Save a setting to database"""
    setting = Settings.query.filter_by(key=key).first()
    if setting:
        setting.value = value
    else:
        setting = Settings(key=key, value=value)
        db.session.add(setting)
    db.session.commit()

def log_activity(campaign_id, action, details, status='success'):
    """Log an activity"""
    log = ActivityLog(
        campaign_id=campaign_id,
        action=action,
        details=details,
        status=status
    )
    db.session.add(log)
    db.session.commit()

# ==================== ROUTES ====================

@app.route('/')
def index():
    """Main dashboard"""
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    """Dashboard page"""
    campaigns = Campaign.query.all()
    total_leads = JobLead.query.count()
    emails_sent = JobLead.query.filter_by(email_sent=True).count()
    recent_logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(10).all()

    stats = {
        'total_campaigns': len(campaigns),
        'active_campaigns': len([c for c in campaigns if c.status == 'active']),
        'total_leads': total_leads,
        'emails_sent': emails_sent
    }

    return render_template('dashboard.html', campaigns=campaigns, stats=stats, logs=recent_logs)

@app.route('/pipeline')
def pipeline_page():
    """Pipeline page - step by step discovery"""
    return render_template('pipeline.html')

@app.route('/leads')
def leads_page():
    """All leads page"""
    return render_template('leads.html')

@app.route('/emails')
def emails_page():
    """Email queue page"""
    return render_template('emails.html')

@app.route('/settings')
def settings_page():
    """Settings page"""
    return render_template('settings.html')

# ==================== API ROUTES ====================

# Settings API
@app.route('/api/settings', methods=['GET'])
def get_settings():
    """Get all settings"""
    settings_dict = {}
    settings = Settings.query.all()
    for setting in settings:
        # Don't expose sensitive keys in full
        if 'key' in setting.key.lower() or 'secret' in setting.key.lower():
            settings_dict[setting.key] = '***' if setting.value else ''
        else:
            settings_dict[setting.key] = setting.value

    return jsonify(settings_dict)

@app.route('/api/settings', methods=['POST'])
def update_settings():
    """Update settings"""
    try:
        data = request.json
        for key, value in data.items():
            # Handle different value types
            if value is None:
                continue

            # For strings, skip masked values
            if isinstance(value, str):
                if value == '***' or not value.strip():
                    continue

            # Save the setting
            save_setting(key, str(value))

        log_activity(None, 'settings_updated', 'Settings were updated', 'success')
        return jsonify({'success': True, 'message': 'Settings updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/settings/test-email', methods=['POST'])
def test_email_config():
    """Test email configuration"""
    try:
        client_id = get_setting('azure_client_id')
        client_secret = get_setting('azure_client_secret')
        tenant_id = get_setting('azure_tenant_id')

        sender = EmailSender(client_id, client_secret, tenant_id)
        result = sender.validate_config()

        return jsonify(result)
    except Exception as e:
        return jsonify({'valid': False, 'message': str(e)})

# Campaign API
@app.route('/api/campaigns', methods=['GET'])
def get_campaigns():
    """Get all campaigns"""
    campaigns = Campaign.query.all()
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'search_keywords': c.search_keywords,
        'company_size_min': c.company_size_min,
        'company_size_max': c.company_size_max,
        'jobs_per_run': c.jobs_per_run,
        'status': c.status,
        'schedule_enabled': c.schedule_enabled,
        'schedule_frequency': c.schedule_frequency,
        'created_at': c.created_at.isoformat()
    } for c in campaigns])

@app.route('/api/campaigns', methods=['POST'])
def create_campaign():
    """Create a new campaign"""
    try:
        data = request.json
        campaign = Campaign(
            name=data['name'],
            search_keywords=data['search_keywords'],
            company_size_min=data.get('company_size_min', 50),
            company_size_max=data.get('company_size_max', 200),
            jobs_per_run=data.get('jobs_per_run', 10),
            email_template_id=data.get('email_template_id'),
            status='draft'
        )

        db.session.add(campaign)
        db.session.commit()

        log_activity(campaign.id, 'campaign_created', f'Campaign "{campaign.name}" created', 'success')

        return jsonify({'success': True, 'campaign_id': campaign.id})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/campaigns/<int:campaign_id>', methods=['PUT'])
def update_campaign(campaign_id):
    """Update a campaign"""
    try:
        campaign = Campaign.query.get_or_404(campaign_id)
        data = request.json

        campaign.name = data.get('name', campaign.name)
        campaign.search_keywords = data.get('search_keywords', campaign.search_keywords)
        campaign.company_size_min = data.get('company_size_min', campaign.company_size_min)
        campaign.company_size_max = data.get('company_size_max', campaign.company_size_max)
        campaign.jobs_per_run = data.get('jobs_per_run', campaign.jobs_per_run)
        campaign.status = data.get('status', campaign.status)
        campaign.schedule_enabled = data.get('schedule_enabled', campaign.schedule_enabled)
        campaign.schedule_frequency = data.get('schedule_frequency', campaign.schedule_frequency)

        db.session.commit()

        # Update scheduler if needed
        if campaign.schedule_enabled:
            campaign_scheduler.schedule_campaign(
                campaign.id,
                campaign.schedule_frequency,
                run_campaign_job
            )
        else:
            campaign_scheduler.remove_campaign(campaign.id)

        log_activity(campaign.id, 'campaign_updated', f'Campaign "{campaign.name}" updated', 'success')

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/campaigns/<int:campaign_id>/run', methods=['POST'])
def run_campaign(campaign_id):
    """Run a campaign"""
    try:
        campaign = Campaign.query.get_or_404(campaign_id)

        # Run campaign in background or synchronously
        result = execute_campaign(campaign)

        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

def execute_campaign(campaign):
    """Execute a campaign - search jobs, find contacts, send emails"""
    try:
        log_activity(campaign.id, 'campaign_started', f'Campaign "{campaign.name}" started', 'success')

        # Get API keys from settings
        google_api_key = get_setting('google_api_key')
        google_cx = get_setting('google_cx_code')
        apollo_api_key = get_setting('apollo_api_key')

        if not all([google_api_key, google_cx, apollo_api_key]):
            return {'success': False, 'message': 'API keys not configured. Please check settings.'}

        # Initialize services
        google_search = GoogleSearchService(google_api_key, google_cx)
        apollo_api = ApolloAPIService(apollo_api_key)

        # Search for jobs
        jobs = google_search.search_jobs(campaign.search_keywords, campaign.jobs_per_run)

        if not jobs:
            log_activity(campaign.id, 'no_jobs_found', 'No jobs found for search criteria', 'warning')
            return {'success': False, 'message': 'No jobs found'}

        processed = 0
        emails_sent = 0

        for job in jobs:
            try:
                # Extract company domain
                domain = google_search.extract_company_from_url(job['link'])
                if not domain:
                    continue

                # Get company info from Apollo
                org_data = apollo_api.search_organization(domain)
                if not org_data:
                    continue

                # Check company size
                if not apollo_api.check_company_size(
                    org_data['estimated_num_employees'],
                    campaign.company_size_min,
                    campaign.company_size_max
                ):
                    continue

                # Find CEO/Owner contacts
                contacts = apollo_api.find_contacts(domain)
                if not contacts:
                    continue

                # Use first contact (highest ranking)
                contact = contacts[0]

                # Generate email
                email_data = email_generator.generate_email(
                    job_data={
                        'job_title': job['title'],
                        'company_name': org_data['name'],
                        'job_url': job['link']
                    },
                    contact_data=contact
                )

                # Save to database
                lead = JobLead(
                    campaign_id=campaign.id,
                    job_title=job['title'],
                    company_name=org_data['name'],
                    company_size=str(org_data['estimated_num_employees']),
                    job_url=job['link'],
                    contact_name=contact['name'],
                    contact_title=contact['title'],
                    contact_email=contact['email'],
                    email_subject=email_data['subject'],
                    email_body=email_data['body'],
                    status='ready'
                )

                db.session.add(lead)
                db.session.commit()

                # Send email if contact email exists
                if contact['email']:
                    sender_email = get_setting('sender_email')
                    if sender_email:
                        sender = EmailSender(
                            get_setting('azure_client_id'),
                            get_setting('azure_client_secret'),
                            get_setting('azure_tenant_id')
                        )

                        send_result = sender.send_email(
                            contact['email'],
                            email_data['subject'],
                            email_data['body'],
                            sender_email
                        )

                        if send_result['success']:
                            lead.email_sent = True
                            lead.email_sent_at = datetime.utcnow()
                            lead.status = 'contacted'
                            db.session.commit()
                            emails_sent += 1

                            # Log to Google Sheets
                            spreadsheet_id = get_setting('google_spreadsheet_id')
                            if spreadsheet_id:
                                sheets_logger = SheetsLogger()
                                sheets_logger.log_job_lead(spreadsheet_id, {
                                    'campaign_name': campaign.name,
                                    'job_title': lead.job_title,
                                    'company_name': lead.company_name,
                                    'company_size': lead.company_size,
                                    'job_url': lead.job_url,
                                    'contact_name': lead.contact_name,
                                    'contact_title': lead.contact_title,
                                    'contact_email': lead.contact_email,
                                    'email_sent': lead.email_sent,
                                    'email_subject': lead.email_subject,
                                    'status': lead.status,
                                    'notes': ''
                                })

                processed += 1

            except Exception as e:
                print(f"Error processing job: {str(e)}")
                continue

        log_activity(
            campaign.id,
            'campaign_completed',
            f'Processed {processed} jobs, sent {emails_sent} emails',
            'success'
        )

        return {
            'success': True,
            'processed': processed,
            'emails_sent': emails_sent,
            'message': f'Campaign completed. Processed {processed} leads, sent {emails_sent} emails.'
        }

    except Exception as e:
        log_activity(campaign.id, 'campaign_failed', str(e), 'error')
        return {'success': False, 'message': str(e)}

def run_campaign_job(campaign_id):
    """Background job to run campaign (called by scheduler)"""
    with app.app_context():
        campaign = Campaign.query.get(campaign_id)
        if campaign:
            execute_campaign(campaign)

# Email Templates API
@app.route('/api/templates', methods=['GET'])
def get_templates():
    """Get all email templates"""
    templates = EmailTemplate.query.all()
    return jsonify([{
        'id': t.id,
        'name': t.name,
        'subject_template': t.subject_template,
        'body_template': t.body_template,
        'is_default': t.is_default
    } for t in templates])

@app.route('/api/templates', methods=['POST'])
def create_template():
    """Create email template"""
    try:
        data = request.json
        template = EmailTemplate(
            name=data['name'],
            subject_template=data['subject_template'],
            body_template=data['body_template'],
            is_default=data.get('is_default', False)
        )

        db.session.add(template)
        db.session.commit()

        return jsonify({'success': True, 'template_id': template.id})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

# Job Leads API
@app.route('/api/leads', methods=['GET'])
def get_leads():
    """Get all job leads with optional filters"""
    campaign_id = request.args.get('campaign_id')
    status = request.args.get('status')
    sent = request.args.get('sent')
    limit = request.args.get('limit', type=int)

    query = JobLead.query

    if campaign_id:
        query = query.filter_by(campaign_id=campaign_id)
    
    if status:
        query = query.filter_by(status=status)
    
    if sent == 'true':
        query = query.filter_by(email_sent=True)
    elif sent == 'false':
        query = query.filter_by(email_sent=False)
    
    query = query.order_by(JobLead.created_at.desc())
    
    if limit:
        query = query.limit(limit)
    
    leads = query.all()

    return jsonify([{
        'id': l.id,
        'campaign_id': l.campaign_id,
        'job_title': l.job_title,
        'company_name': l.company_name,
        'company_size': l.company_size,
        'job_url': l.job_url,
        'contact_name': l.contact_name,
        'contact_title': l.contact_title,
        'contact_email': l.contact_email,
        'email_subject': l.email_subject,
        'email_body': l.email_body,
        'email_sent': l.email_sent,
        'email_sent_at': l.email_sent_at.isoformat() if l.email_sent_at else None,
        'status': l.status,
        'created_at': l.created_at.isoformat()
    } for l in leads])

# Activity Logs API
@app.route('/api/logs', methods=['GET'])
def get_logs():
    """Get activity logs"""
    limit = request.args.get('limit', 50, type=int)
    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(limit).all()

    return jsonify([{
        'id': l.id,
        'campaign_id': l.campaign_id,
        'action': l.action,
        'details': l.details,
        'status': l.status,
        'created_at': l.created_at.isoformat()
    } for l in logs])

# ==================== PIPELINE API ====================

def extract_domain(url):
    """Extract domain from URL"""
    try:
        parsed = urlparse(url)
        domain = parsed.netloc or parsed.path
        # Remove www. prefix
        if domain.startswith('www.'):
            domain = domain[4:]
        # Handle cases like jobs.company.com
        parts = domain.split('.')
        if len(parts) > 2:
            # Return last two parts (company.com)
            return '.'.join(parts[-2:])
        return domain
    except:
        return None

@app.route('/api/pipeline/search', methods=['POST'])
def pipeline_search():
    """Step 1: Search for individual LinkedIn job postings and extract company information"""
    try:
        data = request.json
        keywords = data.get('keywords', '')
        num_results = int(data.get('num_results', 10))

        google_api_key = get_setting('google_api_key')
        google_cx = get_setting('google_cx_code')

        if not google_api_key or not google_cx:
            return jsonify({
                'success': False,
                'message': 'Google API not configured. Go to Settings to add your API keys.'
            })

        google_search = GoogleSearchService(google_api_key, google_cx, use_vector_search=True)
        job_parser = JobParserService(google_api_key, google_cx)

        print(f"\n[SEARCH] Starting LinkedIn job search for: '{keywords}'")
        print(f"[STATS] Requesting {num_results} individual job postings\n")

        # Search for individual LinkedIn job postings only (with vector search enabled)
        search_results = google_search.search_linkedin_jobs(keywords, num_results, use_enhanced_search=True)

        if not search_results:
            return jsonify({
                'success': False,
                'message': 'No LinkedIn jobs found. Try different keywords or check your Google API quota.'
            })

        print(f"\n[OK] Found {len(search_results)} LinkedIn job postings")
        print("="*60)

        # Parse each search result to extract job and company information
        formatted_jobs = []
        seen_companies = set()

        # Import vector search for validation
        from services.vector_search import VectorSearchService
        validator = VectorSearchService() if google_search.use_vector_search else None

        for idx, result in enumerate(search_results, 1):
            print(f"\n[{idx}/{len(search_results)}] Processing: {result.get('title', '')[:60]}...")

            # Parse job data to extract company name and job title
            parsed_job = job_parser.parse_job_data(result)

            company_name = parsed_job['company_name']
            job_title = parsed_job['job_title']

            # Skip if we couldn't extract company name
            if company_name == "Unknown Company":
                print(f"   [WARN] Skipping - could not extract company name")
                continue

            # Validate parsed job title if vector search is enabled
            if validator:
                # Create a validation result with the parsed job title
                validation_result = {
                    'title': job_title,
                    'snippet': parsed_job['job_snippet'],
                    'link': parsed_job['job_link']
                }
                is_valid, quality_score, reason = validator.validate_job_result(validation_result)

                if not is_valid or quality_score < 0.4:
                    print(f"   [ERROR] Invalid job (score: {quality_score:.2f}): {reason}")
                    print(f"      Job title: {job_title[:60]}")
                    continue
                else:
                    print(f"   [OK] Valid job (quality: {quality_score:.2f})")

            # Skip duplicate companies (same company, different job)
            if company_name.lower() in seen_companies:
                print(f"   [WARN] Skipping duplicate company: {company_name}")
                continue

            # Find company domain
            print(f"   [SEARCH] Finding domain for: {company_name}")
            domain = job_parser.find_company_domain(company_name)

            if not domain:
                print(f"   [WARN] Could not find domain for company: {company_name}")
                continue

            seen_companies.add(company_name.lower())

            formatted_jobs.append({
                'title': parsed_job['job_title'],
                'link': result.get('link', ''),
                'snippet': parsed_job['job_snippet'],
                'domain': domain,
                'platform': 'LinkedIn',
                'company_name': company_name
            })

            print(f"   [OK] Added: {parsed_job['job_title']} at {company_name}")
            print(f"   [WEB] Domain: {domain}")

        print("\n" + "="*60)
        print(f"[RESULT] Final Results: {len(formatted_jobs)} unique companies with job openings\n")

        log_activity(None, 'pipeline_search', f'Found {len(formatted_jobs)} LinkedIn jobs for "{keywords}"', 'success')

        return jsonify({
            'success': True,
            'jobs': formatted_jobs,
            'count': len(formatted_jobs)
        })

    except Exception as e:
        print(f"\n[ERROR] Pipeline search error: {str(e)}")
        import traceback
        traceback.print_exc()
        log_activity(None, 'pipeline_search', str(e), 'error')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/pipeline/company', methods=['POST'])
def pipeline_company():
    """Step 2: Get and enrich company details from Apollo"""
    try:
        data = request.json
        domain = data.get('domain')
        min_employees = int(data.get('min_employees', 50))
        max_employees = int(data.get('max_employees', 500))

        apollo_api_key = get_setting('apollo_api_key')

        if not apollo_api_key:
            return jsonify({
                'success': False,
                'message': 'Apollo API not configured. Go to Settings.'
            })

        print(f"\n[COMPANY] Enriching company: {domain}")

        apollo = ApolloAPIService(apollo_api_key)

        # Use enrich_organization to get full company details
        org_data = apollo.enrich_organization(domain)

        if not org_data:
            # Fallback to search if enrich doesn't work
            org_data = apollo.search_organization(domain)

        if not org_data:
            return jsonify({
                'success': False,
                'message': f'Company not found for domain: {domain}'
            })

        # Check company size
        employees = org_data.get('estimated_num_employees', 0)
        if employees < min_employees or employees > max_employees:
            return jsonify({
                'success': False,
                'message': f'Company size ({employees} employees) outside range ({min_employees}-{max_employees})'
            })

        print(f"[OK] Company enriched: {org_data.get('name')} ({employees} employees)")

        return jsonify({
            'success': True,
            'company': {
                # Basic Info
                'id': org_data.get('id', ''),
                'name': org_data.get('name', 'Unknown'),
                'domain': domain,
                'website_url': org_data.get('website_url', f'https://{domain}'),
                'logo_url': org_data.get('logo_url', ''),
                'description': org_data.get('short_description', ''),
                'seo_description': org_data.get('seo_description', ''),
                
                # Size & Industry
                'employees': employees,
                'industry': org_data.get('industry', ''),
                'subindustry': org_data.get('subindustry', ''),
                'founded_year': org_data.get('founded_year', ''),
                'keywords': org_data.get('keywords', []),
                
                # Financial
                'annual_revenue': org_data.get('annual_revenue_printed', ''),
                'total_funding': org_data.get('total_funding_printed', ''),
                'latest_funding_round': org_data.get('latest_funding_round_type', ''),
                'latest_funding_amount': org_data.get('latest_funding_amount', ''),
                'latest_funding_date': org_data.get('latest_funding_date', ''),
                'publicly_traded': org_data.get('publicly_traded_symbol', ''),
                
                # Location
                'city': org_data.get('city', ''),
                'state': org_data.get('state', ''),
                'country': org_data.get('country', ''),
                'address': org_data.get('raw_address', ''),
                'phone': org_data.get('phone', ''),
                
                # Social & Links
                'linkedin': org_data.get('linkedin_url', ''),
                'twitter': org_data.get('twitter_url', ''),
                'facebook': org_data.get('facebook_url', ''),
                'crunchbase': org_data.get('crunchbase_url', ''),
                
                # Tech Stack
                'technologies': org_data.get('technology_names', []),
                
                # Department headcounts
                'department_headcount': org_data.get('departmental_head_count', {})
            }
        })

    except Exception as e:
        print(f"[ERROR] Error enriching company: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/pipeline/contact', methods=['POST'])
def pipeline_contact():
    """Step 3: Find and enrich decision maker contacts with email reveal"""
    try:
        data = request.json
        domain = data.get('domain')
        role_type = data.get('role_type', 'executive')  # executive, tech, hr, sales, marketing, finance, all
        per_page = int(data.get('per_page', 10))
        reveal_all = data.get('reveal_all', False)  # Whether to reveal emails for all contacts

        apollo_api_key = get_setting('apollo_api_key')

        if not apollo_api_key:
            return jsonify({'success': False, 'message': 'Apollo API not configured'})

        print(f"\n[CONTACT] Finding {role_type} decision makers for: {domain}")

        apollo = ApolloAPIService(apollo_api_key)

        # Find contacts using role-based search (without email reveal first to save credits)
        contacts = apollo.find_contacts_by_role(
            domain=domain, 
            role_type=role_type, 
            per_page=per_page,
            reveal_emails=False
        )

        if not contacts:
            return jsonify({
                'success': False,
                'message': f'No {role_type} contacts found'
            })

        # Format all contacts for response
        formatted_contacts = []
        for contact in contacts:
            formatted_contacts.append({
                'id': contact.get('id', ''),
                'name': contact.get('name', 'Unknown'),
                'first_name': contact.get('first_name', ''),
                'last_name': contact.get('last_name', ''),
                'title': contact.get('title', ''),
                'role_category': contact.get('role_category', 'Other'),
                'email': contact.get('email', ''),
                'email_status': contact.get('email_status', ''),
                'phone_numbers': contact.get('phone_numbers', []),
                'linkedin': contact.get('linkedin_url', ''),
                'twitter': contact.get('twitter_url', ''),
                'company': contact.get('organization_name', ''),
                'city': contact.get('city', ''),
                'state': contact.get('state', ''),
                'country': contact.get('country', ''),
                'seniority': contact.get('seniority', ''),
                'departments': contact.get('departments', []),
                'photo_url': contact.get('photo_url', ''),
                'email_revealed': False
            })

        # Get the primary contact (first one, usually highest ranking)
        primary_contact = formatted_contacts[0]

        print(f"[OK] Found {len(formatted_contacts)} contacts, primary: {primary_contact.get('name')} - {primary_contact.get('title')}")
        print(f"[LOCK] Unlocking email for primary contact...")

        # Reveal email for primary contact
        enriched_contact = apollo.enrich_person(
            person_id=primary_contact.get('id'),
            domain=domain,  # Pass domain for guessed emails
            reveal_emails=True
        )

        if enriched_contact:
            if enriched_contact.get('email'):
                primary_contact['email'] = enriched_contact.get('email')
                primary_contact['email_status'] = enriched_contact.get('email_status', 'verified')
                primary_contact['email_revealed'] = True
                print(f"[OK] Email unlocked: {primary_contact['email']}")
            else:
                # No email from Apollo - include status and guessed emails
                primary_contact['email_status'] = enriched_contact.get('email_status', 'unavailable')
                primary_contact['email_status_explanation'] = enriched_contact.get('email_status_explanation', '')
                primary_contact['guessed_emails'] = enriched_contact.get('guessed_emails', [])
                print(f"[WARN] No email from Apollo - status: {primary_contact['email_status']}")
                if primary_contact.get('guessed_emails'):
                    print(f"[TIP] Suggested emails: {', '.join(primary_contact['guessed_emails'][:3])}")
            
            formatted_contacts[0] = primary_contact
        
        # If reveal_all is True, reveal emails for all contacts
        if reveal_all and len(formatted_contacts) > 1:
            print(f"\n[LOCK] Revealing emails for remaining {len(formatted_contacts) - 1} contacts...")
            for i, contact in enumerate(formatted_contacts[1:], 1):
                enriched = apollo.enrich_person(
                    person_id=contact.get('id'),
                    domain=domain,
                    reveal_emails=True
                )
                if enriched:
                    if enriched.get('email'):
                        formatted_contacts[i]['email'] = enriched.get('email')
                        formatted_contacts[i]['email_status'] = enriched.get('email_status', 'verified')
                        formatted_contacts[i]['email_revealed'] = True
                    else:
                        formatted_contacts[i]['email_status'] = enriched.get('email_status', 'unavailable')
                        formatted_contacts[i]['guessed_emails'] = enriched.get('guessed_emails', [])

        return jsonify({
            'success': True,
            'contact': primary_contact,
            'all_contacts': formatted_contacts,
            'total_found': len(formatted_contacts),
            'role_type': role_type
        })

    except Exception as e:
        print(f"[ERROR] Error finding contact: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/pipeline/reveal-email', methods=['POST'])
def pipeline_reveal_email():
    """Step 3.5: Reveal contact email using Apollo /v1/people/match (uses credits)"""
    try:
        data = request.json
        person_id = data.get('person_id')

        apollo_api_key = get_setting('apollo_api_key')

        if not apollo_api_key:
            return jsonify({'success': False, 'message': 'Apollo API not configured'})

        if not person_id:
            return jsonify({'success': False, 'message': 'Person ID required'})

        print(f"\n[LOCK] Manual email reveal requested for person ID: {person_id}")

        apollo = ApolloAPIService(apollo_api_key)

        # Use enrich_person which calls /v1/people/match
        enriched = apollo.enrich_person(person_id=person_id, reveal_emails=True)

        if enriched and enriched.get('email'):
            email = enriched['email']
            log_activity(None, 'email_revealed', f'Email revealed via Apollo: {email}', 'success')

            return jsonify({
                'success': True,
                'email': email,
                'email_status': enriched.get('email_status', 'unknown'),
                'full_contact': enriched
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Apollo did not return an email. Check console logs for details.'
            })

    except Exception as e:
        print(f"[ERROR] Error in reveal-email endpoint: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/pipeline/generate-email', methods=['POST'])
def pipeline_generate_email():
    """Step 4: Generate personalized email"""
    try:
        data = request.json
        job = data.get('job', {})
        company = data.get('company', {})
        contact = data.get('contact', {})
        
        # Generate email using AI
        email_data = email_generator.generate_email(
            job_data={
                'job_title': job.get('title', 'position'),
                'company_name': company.get('name', 'your company'),
                'job_url': job.get('link', '')
            },
            contact_data=contact
        )
        
        # Save as a new lead
        lead = JobLead(
            job_title=job.get('title'),
            company_name=company.get('name'),
            company_size=str(company.get('employees', '')),
            job_url=job.get('link'),
            contact_name=contact.get('name'),
            contact_title=contact.get('title'),
            contact_email=contact.get('email'),
            email_subject=email_data['subject'],
            email_body=email_data['body'],
            status='ready'
        )
        
        db.session.add(lead)
        db.session.commit()
        
        log_activity(None, 'lead_created', f'Lead created for {company.get("name")}', 'success')
        
        return jsonify({
            'success': True,
            'lead_id': lead.id,
            'email': email_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# ==================== ENHANCED LEADS API ====================

@app.route('/api/leads/<int:lead_id>', methods=['GET'])
def get_lead_detail(lead_id):
    """Get single lead details"""
    lead = JobLead.query.get_or_404(lead_id)
    return jsonify({
        'id': lead.id,
        'campaign_id': lead.campaign_id,
        'job_title': lead.job_title,
        'company_name': lead.company_name,
        'company_size': lead.company_size,
        'job_url': lead.job_url,
        'contact_name': lead.contact_name,
        'contact_title': lead.contact_title,
        'contact_email': lead.contact_email,
        'email_subject': lead.email_subject,
        'email_body': lead.email_body,
        'email_sent': lead.email_sent,
        'email_sent_at': lead.email_sent_at.isoformat() if lead.email_sent_at else None,
        'status': lead.status,
        'notes': lead.notes,
        'created_at': lead.created_at.isoformat(),
        'updated_at': lead.updated_at.isoformat() if lead.updated_at else None
    })

@app.route('/api/leads/<int:lead_id>', methods=['PUT'])
def update_lead(lead_id):
    """Update lead details"""
    try:
        lead = JobLead.query.get_or_404(lead_id)
        data = request.json
        
        if 'email_subject' in data:
            lead.email_subject = data['email_subject']
        if 'email_body' in data:
            lead.email_body = data['email_body']
        if 'status' in data:
            lead.status = data['status']
        if 'notes' in data:
            lead.notes = data['notes']
        
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/leads/<int:lead_id>/send', methods=['POST'])
def send_lead_email(lead_id):
    """Send email for a specific lead"""
    try:
        lead = JobLead.query.get_or_404(lead_id)
        data = request.json or {}
        
        # Allow overriding subject/body
        subject = data.get('subject', lead.email_subject)
        body = data.get('body', lead.email_body)
        
        if not lead.contact_email:
            return jsonify({'success': False, 'message': 'No email address for this contact'})
        
        # Get email sender settings
        client_id = get_setting('azure_client_id')
        client_secret = get_setting('azure_client_secret')
        tenant_id = get_setting('azure_tenant_id')
        sender_email = get_setting('sender_email')
        
        if not all([client_id, client_secret, tenant_id, sender_email]):
            return jsonify({'success': False, 'message': 'Email settings not configured. Go to Settings.'})
        
        sender = EmailSender(client_id, client_secret, tenant_id)
        result = sender.send_email(
            lead.contact_email,
            subject,
            body,
            sender_email
        )
        
        if result.get('success'):
            lead.email_sent = True
            lead.email_sent_at = datetime.utcnow()
            lead.email_subject = subject
            lead.email_body = body
            lead.status = 'contacted'
            db.session.commit()
            
            log_activity(None, 'email_sent', f'Email sent to {lead.contact_name} at {lead.company_name}', 'success')
            
            return jsonify({'success': True, 'message': 'Email sent successfully'})
        else:
            return jsonify({'success': False, 'message': result.get('message', 'Failed to send email')})
        
    except Exception as e:
        log_activity(None, 'email_failed', str(e), 'error')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/leads/<int:lead_id>/skip', methods=['POST'])
def skip_lead(lead_id):
    """Skip a lead"""
    try:
        lead = JobLead.query.get_or_404(lead_id)
        lead.status = 'skipped'
        db.session.commit()

        log_activity(None, 'lead_skipped', f'Skipped lead: {lead.company_name}', 'info')

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# ==================== AI FEATURES API ====================

@app.route('/api/ai/score-lead', methods=['POST'])
def ai_score_lead():
    """Score a single lead using AI"""
    try:
        data = request.json
        lead_data = data.get('lead', {})
        query = data.get('query', '')

        scoring = ai_lead_scorer.score_lead(lead_data, query)

        return jsonify({
            'success': True,
            'scoring': scoring
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/ai/score-leads-batch', methods=['POST'])
def ai_score_leads_batch():
    """Score multiple leads at once"""
    try:
        data = request.json
        leads = data.get('leads', [])
        query = data.get('query', '')

        scored_leads = ai_lead_scorer.batch_score_leads(leads, query)

        return jsonify({
            'success': True,
            'leads': scored_leads,
            'distribution': ai_lead_scorer.get_lead_priority_distribution(scored_leads)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/ai/research-company', methods=['POST'])
def ai_research_company():
    """AI-powered company research"""
    try:
        from services.ai_research_agent import AIResearchAgent

        data = request.json
        company_name = data.get('company_name', '')
        company_domain = data.get('company_domain', '')
        company_data = data.get('company_data', {})
        mode = data.get('mode', 'quick')  # 'quick' or 'full'

        google_api_key = get_setting('google_api_key')
        google_cx = get_setting('google_cx_code')

        researcher = AIResearchAgent(google_api_key, google_cx)

        if mode == 'quick':
            research = researcher.quick_research(company_name, company_data)
        else:
            research = researcher.research_company(company_name, company_domain, company_data)

        return jsonify({
            'success': True,
            'research': research
        })
    except Exception as e:
        print(f"Research error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/analytics/dashboard', methods=['GET'])
def get_dashboard_analytics():
    """Get enhanced analytics for dashboard"""
    try:
        # Get all leads
        all_leads = JobLead.query.all()
        campaigns = Campaign.query.all()

        # Basic stats
        total_leads = len(all_leads)
        emails_sent = sum(1 for l in all_leads if l.email_sent)

        # Response rate (mock for now - would need reply tracking)
        response_rate = 0.15  # 15% placeholder

        # Leads by status
        status_distribution = {}
        for lead in all_leads:
            status = lead.status or 'unknown'
            status_distribution[status] = status_distribution.get(status, 0) + 1

        # Recent activity trend (last 7 days)
        from datetime import timedelta
        today = datetime.utcnow()
        week_ago = today - timedelta(days=7)

        recent_leads = [l for l in all_leads if l.created_at >= week_ago]

        # Group by date
        daily_stats = {}
        for i in range(7):
            date = (today - timedelta(days=i)).date()
            daily_stats[date.isoformat()] = {
                'leads': 0,
                'emails': 0
            }

        for lead in recent_leads:
            date_key = lead.created_at.date().isoformat()
            if date_key in daily_stats:
                daily_stats[date_key]['leads'] += 1
                if lead.email_sent:
                    daily_stats[date_key]['emails'] += 1

        analytics = {
            'overview': {
                'total_campaigns': len(campaigns),
                'active_campaigns': sum(1 for c in campaigns if c.status == 'active'),
                'total_leads': total_leads,
                'emails_sent': emails_sent,
                'response_rate': response_rate,
                'conversion_rate': (response_rate * 0.3)  # Mock 30% of responses convert
            },
            'status_distribution': status_distribution,
            'daily_trend': daily_stats,
            'top_performers': {
                'best_campaign': campaigns[0].name if campaigns else 'None',
                'most_responsive_industry': 'Technology'  # Mock
            }
        }

        return jsonify({
            'success': True,
            'analytics': analytics
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# ==================== MANUFACTURING ICP API ====================

@app.route('/manufacturing-icp')
def manufacturing_icp_page():
    """Manufacturing ICP lead generator page"""
    return render_template('manufacturing_icp.html')

@app.route('/api/manufacturing-icp/campaigns', methods=['GET'])
def get_manufacturing_campaigns():
    """Get all Manufacturing ICP campaigns"""
    campaigns = ManufacturingICPCampaign.query.order_by(ManufacturingICPCampaign.created_at.desc()).all()
    return jsonify([c.to_dict() for c in campaigns])

@app.route('/api/manufacturing-icp/generate', methods=['POST'])
def generate_manufacturing_leads():
    """Generate Manufacturing ICP leads with custom filters"""
    try:
        data = request.json

        # Get Apollo API key
        apollo_api_key = get_setting('apollo_api_key')
        if not apollo_api_key:
            return jsonify({
                'success': False,
                'message': 'Apollo API key not configured. Go to Settings.'
            })

        # Extract filters from request (can be at root or inside 'filters' object)
        filters_data = data.get('filters', data)  # Support both formats

        # Create campaign
        campaign = ManufacturingICPCampaign(
            name=data.get('campaign_name', 'Manufacturing ICP'),
            t1_target=int(data.get('t1_target', data.get('t1_count', 20))),
            t2_target=int(data.get('t2_target', data.get('t2_count', 20))),
            t3_target=int(data.get('t3_target', data.get('t3_count', 10))),
            industries=json.dumps(filters_data.get('industries', [])),
            t1_titles=json.dumps(filters_data.get('t1_titles', [])),
            t2_titles=json.dumps(filters_data.get('t2_titles', [])),
            t3_titles=json.dumps(filters_data.get('t3_titles', [])),
            locations=json.dumps(filters_data.get('locations', {'usa': True, 'india': True})),
            size_min=int(filters_data.get('size_min', 200)),
            size_max=int(filters_data.get('size_max', 10000)),
            min_validation_score=int(filters_data.get('min_validation_score', filters_data.get('min_score', 4))),
            status='in_progress'
        )

        db.session.add(campaign)
        db.session.commit()

        log_activity(None, 'manufacturing_icp_started', f'Started Manufacturing ICP campaign: {campaign.name}', 'success')

        # Generate leads with RAG-powered intelligence
        from services.manufacturing_icp import ManufacturingICPService

        print(f"[DEBUG] Initializing Manufacturing ICP service...")
        print(f"[DEBUG] Apollo API key: {apollo_api_key[:10]}...")
        
        # Initialize service with RAG ENABLED for 10-20x speedup
        try:
            service = ManufacturingICPService(
                apollo_api_key,
                use_rag=True,  # Enable RAG for semantic pre-filtering
                use_ollama=False  # Use sentence-transformers (faster, no Ollama needed)
            )
            print(f"[DEBUG] Service initialized successfully with RAG intelligence")
        except Exception as e:
            print(f"[ERROR] Failed to initialize service: {e}")
            raise

        filters = {
            'industries': json.loads(campaign.industries) if campaign.industries else None,
            # Don't include empty title lists - let service use its defaults
            'locations': json.loads(campaign.locations) if campaign.locations else None,
            'size_min': campaign.size_min,
            'size_max': campaign.size_max,
            'min_validation_score': campaign.min_validation_score
        }
        
        # Only add title filters if they are non-empty arrays
        t1_titles = json.loads(campaign.t1_titles) if campaign.t1_titles else []
        if t1_titles:
            filters['t1_titles'] = t1_titles
            
        t2_titles = json.loads(campaign.t2_titles) if campaign.t2_titles else []
        if t2_titles:
            filters['t2_titles'] = t2_titles
            
        t3_titles = json.loads(campaign.t3_titles) if campaign.t3_titles else []
        if t3_titles:
            filters['t3_titles'] = t3_titles
        
        print(f"[DEBUG] Filters: industries={len(filters.get('industries') or [])}, locations={filters.get('locations')}")
        print(f"[DEBUG] Size range: {filters['size_min']}-{filters['size_max']}")
        print(f"[DEBUG] Calling generate_leads...")

        results = service.generate_leads(campaign, filters)
        
        print(f"[DEBUG] Lead generation complete!")
        print(f"[DEBUG] Results: T1={len(results['t1_leads'])}, T2={len(results['t2_leads'])}, T3={len(results['t3_leads'])}")

        # Save leads to database
        total_saved = 0
        for lead_data in results['t1_leads'] + results['t2_leads'] + results['t3_leads']:
            lead = ManufacturingLead(
                campaign_id=campaign.id,
                tier=lead_data['tier'],
                company_name=lead_data['company']['name'],
                company_domain=lead_data['company']['domain'],
                company_size=lead_data['company']['size'],
                company_industry=lead_data['company']['industry'],
                company_location=lead_data['company']['location'],
                company_revenue=lead_data['company']['revenue'],
                company_linkedin=lead_data['company']['linkedin'],
                company_website=lead_data['company']['website'],
                contact_name=lead_data['contact']['name'],
                contact_title=lead_data['contact']['title'],
                contact_email=lead_data['contact']['email'],
                contact_phone=lead_data['contact']['phone'],
                contact_linkedin=lead_data['contact']['linkedin'],
                email_status=lead_data['contact']['email_status'],
                validation_score=lead_data['validation']['score'],
                validation_details=json.dumps(lead_data['validation']),
                status='new'
            )
            db.session.add(lead)
            total_saved += 1

        # Update campaign status
        campaign.status = 'completed'
        campaign.total_leads = total_saved
        campaign.t1_generated = len(results['t1_leads'])
        campaign.t2_generated = len(results['t2_leads'])
        campaign.t3_generated = len(results['t3_leads'])
        campaign.avg_validation_score = results['avg_score']
        campaign.completed_at = datetime.utcnow()

        db.session.commit()

        log_activity(None, 'manufacturing_icp_completed',
                    f'Completed Manufacturing ICP campaign: {campaign.name}. Generated {total_saved} leads.',
                    'success')

        # Get all saved leads for return
        all_leads = ManufacturingLead.query.filter_by(campaign_id=campaign.id).all()
        leads_data = [lead.to_dict() for lead in all_leads]

        return jsonify({
            'success': True,
            'campaign_id': campaign.id,
            'total_leads': total_saved,
            't1_count': len(results['t1_leads']),
            't2_count': len(results['t2_leads']),
            't3_count': len(results['t3_leads']),
            'avg_score': results['avg_score'],
            'leads': leads_data,
            'summary': {
                'total_generated': total_saved,
                'avg_score': results['avg_score']
            }
        })

    except Exception as e:
        print(f"[ERROR] Manufacturing ICP generation failed: {str(e)}")
        import traceback
        traceback.print_exc()

        if 'campaign' in locals():
            campaign.status = 'failed'
            db.session.commit()

        log_activity(None, 'manufacturing_icp_failed', str(e), 'error')
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/manufacturing-icp/leads/<int:campaign_id>', methods=['GET'])
def get_manufacturing_leads(campaign_id):
    """Get all leads for a Manufacturing ICP campaign"""
    campaign = ManufacturingICPCampaign.query.get_or_404(campaign_id)
    leads = ManufacturingLead.query.filter_by(campaign_id=campaign_id).all()
    
    return jsonify({
        'success': True,
        'campaign': campaign.to_dict(),
        'leads': [lead.to_dict() for lead in leads]
    })

@app.route('/api/manufacturing-icp/export/<int:campaign_id>', methods=['GET'])
def export_manufacturing_leads(campaign_id):
    """Export Manufacturing ICP leads to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        import io

        campaign = ManufacturingICPCampaign.query.get_or_404(campaign_id)
        leads = ManufacturingLead.query.filter_by(campaign_id=campaign_id).all()

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(['Campaign Name', campaign.name])
        ws_summary.append(['Total Leads', campaign.total_leads])
        ws_summary.append(['T1 Leads', campaign.t1_generated])
        ws_summary.append(['T2 Leads', campaign.t2_generated])
        ws_summary.append(['T3 Leads', campaign.t3_generated])
        ws_summary.append(['Avg Validation Score', f"{campaign.avg_validation_score}/6"])
        ws_summary.append(['Date Generated', campaign.completed_at.strftime('%Y-%m-%d %H:%M') if campaign.completed_at else 'N/A'])

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Helper function to create tier sheet
        def create_tier_sheet(tier_name, tier_leads):
            ws = wb.create_sheet(tier_name)
            headers = ['Company', 'Contact Name', 'Title', 'Email', 'Phone', 'Size', 'Industry',
                      'Location', 'Revenue', 'LinkedIn', 'Website', 'Validation Score', 'Checklist']
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add data
            for lead in tier_leads:
                validation = json.loads(lead.validation_details) if lead.validation_details else {}
                checklist_summary = f"{lead.validation_score}/6"

                ws.append([
                    lead.company_name,
                    lead.contact_name,
                    lead.contact_title,
                    lead.contact_email,
                    lead.contact_phone,
                    lead.company_size,
                    lead.company_industry,
                    lead.company_location,
                    lead.company_revenue,
                    lead.contact_linkedin,
                    lead.company_website,
                    checklist_summary,
                    ', '.join([k for k, v in validation.get('checklist', {}).items() if v.get('passed')])
                ])

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Create sheets for each tier
        t1_leads = [l for l in leads if l.tier == 'T1']
        t2_leads = [l for l in leads if l.tier == 'T2']
        t3_leads = [l for l in leads if l.tier == 'T3']

        if t1_leads:
            create_tier_sheet('T1_Decision_Makers', t1_leads)
        if t2_leads:
            create_tier_sheet('T2_HR_Leaders', t2_leads)
        if t3_leads:
            create_tier_sheet('T3_HR_Practitioners', t3_leads)

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"Manufacturing_ICP_{campaign.name.replace(' ', '_')}_{campaign.total_leads}leads.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"[ERROR] Export failed: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

# Initialize database and create default data
def init_db():
    """Initialize database with default data"""
    with app.app_context():
        db.create_all()

        # Create default email template if none exists
        if EmailTemplate.query.count() == 0:
            default_template = EmailTemplate(
                name='Default Professional Template',
                subject_template='Top Talent Available for {job_title} at {company_name}',
                body_template='''Dear {contact_name},

I hope this email finds you well. I noticed that {company_name} is currently hiring for a {job_title} position, and I wanted to reach out.

We are a specialized recruitment agency with a pool of pre-vetted, highly qualified candidates who are actively seeking opportunities in this field. Our candidates have been thoroughly screened and are ready for immediate placement.

I'd love to discuss how we can support your hiring needs and potentially save you time in finding the right talent for this role.

Would you be open to a brief conversation this week?

Best regards,
Recruitment Team''',
                is_default=True
            )
            db.session.add(default_template)
            db.session.commit()

        print("Database initialized successfully!")

if __name__ == '__main__':
    init_db()

    # Load AI model in background
    print("Loading AI model...")
    email_generator.load_model()

    # Run Flask app
    app.run(debug=True, host='0.0.0.0', port=5000)
